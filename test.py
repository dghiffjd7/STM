import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from openpyxl import load_workbook, Workbook

def select_folder():
    path = filedialog.askdirectory(title="選擇來源資料夾")
    if path:
        src_folder_var.set(path)

def select_target_existing():
    file = filedialog.askopenfilename(title="選擇目標 Excel 檔案", filetypes=[("Excel files","*.xlsx")])
    if file:
        target_file_var.set(file)

def select_target_new():
    file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")], title="建立新的目標 Excel 檔案")
    if file:
        target_file_var.set(file)

def gather_excel_files(folder):
    fs = []
    for name in os.listdir(folder):
        if name.startswith("~$"):
            continue
        p = os.path.join(folder, name)
        if os.path.isfile(p) and name.lower().endswith((".xlsx",".xls")):
            fs.append(p)
    return sorted(fs)

def read_ab_from_first_sheet(path):
    try:
        df = pd.read_excel(path, sheet_name=0, header=None, usecols=[0,1], engine=None)
    except Exception:
        df = pd.read_excel(path, sheet_name=0, header=None, usecols=[0,1])
    if df.shape[0] <= 1:
        return pd.DataFrame(columns=[0,1])
    df = df.iloc[1:].copy()
    df = df.loc[~(df[0].isna() & df[1].isna())]
    df.reset_index(drop=True, inplace=True)
    return df

def write_overwrite_ab(tgt_path, data):
    if os.path.exists(tgt_path):
        wb = load_workbook(tgt_path)
    else:
        wb = Workbook()
    ws = wb.worksheets[0]
    maxr = ws.max_row
    for r in range(1, maxr+1):
        ws.cell(row=r, column=1, value=None)
        ws.cell(row=r, column=2, value=None)
    for i in range(len(data)):
        ws.cell(row=i+1, column=1, value=None if pd.isna(data.iat[i,0]) else data.iat[i,0])
        ws.cell(row=i+1, column=2, value=None if pd.isna(data.iat[i,1]) else data.iat[i,1])
    wb.save(tgt_path)

def run():
    try:
        folder = src_folder_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("錯誤","請選擇來源資料夾")
            return
        files = gather_excel_files(folder)
        if not files:
            messagebox.showerror("錯誤","資料夾內無可處理的 Excel 檔")
            return
        combined = []
        for f in files:
            try:
                df = read_ab_from_first_sheet(f)
                if not df.empty:
                    combined.append(df)
            except Exception as e:
                pass
        if not combined:
            messagebox.showerror("錯誤","沒有可合併的資料")
            return
        merged = pd.concat(combined, axis=0, ignore_index=True)
        tgt = target_file_var.get().strip()
        if tgt:
            if os.path.exists(tgt):
                ok = messagebox.askokcancel("覆蓋警告","將覆蓋目標檔案 A、B 欄，是否繼續？")
                if not ok:
                    return
                write_overwrite_ab(tgt, merged)
                messagebox.showinfo("完成", f"已覆蓋寫入 {len(merged)} 筆至 {tgt}")
            else:
                write_overwrite_ab(tgt, merged)
                messagebox.showinfo("完成", f"已建立並寫入 {len(merged)} 筆至 {tgt}")
        else:
            ask = messagebox.askyesno("建立目標檔","未選擇既有檔案，是否建立新檔？")
            if not ask:
                return
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")], title="另存新檔")
            if not path:
                return
            target_file_var.set(path)
            write_overwrite_ab(path, merged)
            messagebox.showinfo("完成", f"已建立並寫入 {len(merged)} 筆至 {path}")
    except Exception as e:
        messagebox.showerror("錯誤", str(e))

root = tk.Tk()
root.title("資料夾批次合併 A,B 欄")
frm = tk.Frame(root, padx=10, pady=10)
frm.pack(fill="both", expand=True)

src_folder_var = tk.StringVar()
target_file_var = tk.StringVar()

tk.Label(frm, text="來源資料夾：").grid(row=0, column=0, sticky="w")
row0 = tk.Frame(frm)
row0.grid(row=0, column=1, sticky="w")
tk.Button(row0, text="選擇資料夾", command=select_folder).pack(side="left", padx=2)
tk.Entry(frm, textvariable=src_folder_var, width=60).grid(row=1, column=0, columnspan=2, sticky="we", pady=4)

tk.Label(frm, text="目標檔案：").grid(row=2, column=0, sticky="w")
row2 = tk.Frame(frm)
row2.grid(row=2, column=1, sticky="w")
tk.Button(row2, text="選擇既有", command=select_target_existing).pack(side="left", padx=2)
tk.Button(row2, text="建立新檔", command=select_target_new).pack(side="left", padx=2)
tk.Entry(frm, textvariable=target_file_var, width=60).grid(row=3, column=0, columnspan=2, sticky="we", pady=4)

tk.Button(frm, text="執行合併", command=run).grid(row=4, column=0, columnspan=2, pady=10)

root.mainloop()
